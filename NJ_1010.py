# -*- coding: utf-8 -*-
"""
Created on Mon Sep 11 17:09:36 2017

@author: MGuidry
"""

import httplib2
import json
from openpyxl import Workbook
import csv

#f=open('test.pkl','rb')
#votes_dict=pickle.load(f)
#f.close()

h = httplib2.Http(ca_certs='./data_files/cacerts.txt')

url_state='https://api.decisiondeskhq.com/voting/1010/'        
    
url='https://api.decisiondeskhq.com/voting/county/1010/'
method = 'GET'
headers = {
    'Accept': 'application/json',
    'Authorization': 'Bearer d97d447fbf3aa810dabdeced4e4b1cefcfff1d17',
    'Content-Type': 'application/json; charset=UTF-8'
}
body=''

response, content = h.request(
        url,
        method,
        body,
        headers)
        

 
votes_dict=dict()

d=json.loads(content)[0]
details=d['Governor']['details']
race=details[0][0]
state=race.keys()[0]
counties=race[state]['counties'].keys()

for county in counties:
    party=race[state]['counties'][county]['candidates'][0]['candidate']['party']
    if county not in votes_dict:
        votes_dict[county]=dict()
    cands=race[state]['counties'][county]['candidates']
    for cand in cands:
        cand_dict=cand['candidate']
        lname=cand_dict['lname']                
        votes_dict[county][lname]=int(cand_dict['votes'])
              #print cand_dict['precincts']

wb=Workbook()
ws = wb.create_sheet("Results")

k=1
for county in sorted(votes_dict.keys()):
    for cand in sorted(votes_dict[county].keys()):
        ws.cell(row=k,column=1).value=county
        ws.cell(row=k,column=2).value=cand
        ws.cell(row=k,column=3).value=votes_dict[county][cand]
        k=k+1

wb.save('1010.xlsx')

#m=open('./data_files/2017_Moore_margin.csv','wb')
#s=open('./data_files/2017_Strange_margin.csv','wb')
#d=open('./data_files/2017_Moore_vs_Strange_margin.csv','wb')
#m_writer=csv.writer(m)
#s_writer=csv.writer(s)
#d_writer=csv.writer(d)
#for county in sorted(votes_dict.keys()):
#    total=float(sum(votes_dict[county].values()))
#    m_margin=votes_dict[county]['Moore']/total
#    s_margin=votes_dict[county]['Strange']/total
#    d_margin=m_margin-s_margin
#    m_writer.writerow([county,m_margin])
#    s_writer.writerow([county,s_margin])
#    d_writer.writerow([county,d_margin])
#
#m.close()
#s.close()
#d.close()
